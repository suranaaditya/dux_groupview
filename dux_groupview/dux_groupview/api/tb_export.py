"""Trial Balance xlsx export.

Streams the full pivot dataset (the same view the cockpit renders) as a
well-formatted, indented Excel workbook. Reuses the pivot reader for
its snapshot fetch + bubble-up aggregation so the export and the
on-screen pivot can never disagree.

Sign convention: applies the standard `FLIP_ROOT_TYPES` natural-side
flip per account, matching every other cockpit reader. Zero-row hiding
mirrors the pivot's frontend rule -- an account is included only if it
(or any descendant) has a non-zero balance in at least one in-scope
company.
"""

import re
from datetime import datetime
from io import BytesIO

import frappe
from frappe import _
from frappe.utils import flt, getdate

from dux_groupview.dux_groupview.api.pivot import (
	_build_accounts_and_balances,
	_build_trusts,
	_require_cockpit_role,
	_resolve_scope,
	_sql_in_tuple,
)
from dux_groupview.dux_groupview.api.utils import FLIP_ROOT_TYPES


# Indian-grouped number format with parens for negatives and a dash for
# zero. `#,##,##0.00` is the lakh/crore grouping pattern Excel renders
# natively when the format string carries the second comma.
_NUM_FMT = '#,##,##0.00;(#,##,##0.00);"-"'


@frappe.whitelist()
def export_tb_xlsx(snapshot_date, companies=None):
	"""Stream a styled .xlsx of the trial balance for one snapshot date.

	`companies` follows the same intersection-with-User-Permissions rule
	as `get_pivot_data` -- a user cannot widen visibility through this
	parameter.
	"""
	_require_cockpit_role()
	snapshot_date = getdate(snapshot_date)

	allowed = _resolve_scope(companies)
	if not allowed:
		_send_xlsx(_build_empty_xlsx(snapshot_date), snapshot_date)
		return

	rows = frappe.db.sql(
		"""
		SELECT r.company, r.account, r.balance, r.root_type, r.account_type,
		       a.account_name, a.is_group, a.parent_account
		FROM `tabDGV TB Snapshot Row` r
		LEFT JOIN `tabAccount` a ON a.name = r.account
		WHERE r.snapshot_date = %(snapshot_date)s
		  AND r.company IN %(companies)s
		""",
		{
			"snapshot_date": snapshot_date,
			"companies": _sql_in_tuple(allowed),
		},
		as_dict=True,
	)

	trusts = _build_trusts(allowed)
	accounts, balances = _build_accounts_and_balances(rows)

	xlsx_bytes = _build_tb_xlsx(snapshot_date, trusts, accounts, balances)
	_send_xlsx(xlsx_bytes, snapshot_date)


# ---------------------------------------------------------------------------
# Builder
# ---------------------------------------------------------------------------

def _build_tb_xlsx(snapshot_date, trusts, accounts, balances):
	"""Construct the workbook and return its bytes."""
	from openpyxl import Workbook
	from openpyxl.styles import Alignment, Font, PatternFill
	from openpyxl.utils import get_column_letter

	# Flatten companies in trust order so columns mirror the pivot UI.
	company_cols = []  # [(trust_id, trust_name, trust_color, company)]
	for t in trusts:
		for c in t["companies"]:
			company_cols.append((t["id"], t["name"], t["color"], c))
	n_cos = len(company_cols)
	total_col = 2 + n_cos  # 1-based; last column is row total

	wb = Workbook()
	ws = wb.active
	ws.title = "Trial Balance"

	header_font = Font(bold=True, color="FFFFFF")
	group_font = Font(bold=True)
	left_align = Alignment(horizontal="left", vertical="center")
	right_align = Alignment(horizontal="right", vertical="center")
	wrap_center = Alignment(
		horizontal="center", vertical="center", wrap_text=True
	)
	# Account labels use this much indent per depth level. 2 indent
	# units ≈ 6 character widths in Excel -- a comfortable visual step
	# per hierarchy level. Numeric cells mirror the same step on the
	# right side (`indent` on a right-aligned cell pulls the value
	# leftward), so each row visually steps inward across every column.
	LABEL_INDENT_PER_DEPTH = 2
	NUMBER_INDENT_PER_DEPTH = 2

	# --- Row 1: title ---
	ws.cell(row=1, column=1, value=f"Trial Balance — {snapshot_date.isoformat()}")
	ws.cell(row=1, column=1).font = Font(bold=True, size=14)
	ws.cell(row=1, column=1).alignment = left_align
	ws.merge_cells(
		start_row=1, end_row=1, start_column=1, end_column=total_col,
	)
	ws.row_dimensions[1].height = 22

	# --- Row 2: trust header (merged across each trust's company span) ---
	acct_header = ws.cell(row=2, column=1, value="Account")
	acct_header.font = group_font
	acct_header.alignment = left_align
	acct_header.fill = PatternFill("solid", fgColor="F1F3F5")

	col = 2
	for t in trusts:
		span = len(t["companies"])
		if span == 0:
			continue
		cell = ws.cell(row=2, column=col, value=t["name"])
		cell.font = header_font
		cell.fill = PatternFill(
			"solid", fgColor=_hex_to_argb(t["color"]),
		)
		cell.alignment = wrap_center
		if span > 1:
			ws.merge_cells(
				start_row=2, end_row=2,
				start_column=col, end_column=col + span - 1,
			)
		col += span

	total_header = ws.cell(row=2, column=total_col, value="Group Total")
	total_header.font = group_font
	total_header.alignment = wrap_center
	total_header.fill = PatternFill("solid", fgColor="F1F3F5")

	# --- Row 3: company header ---
	for i, (_tid, _tname, _tcolor, company) in enumerate(company_cols):
		cell = ws.cell(row=3, column=2 + i, value=company)
		cell.font = Font(bold=True)
		cell.alignment = wrap_center
		cell.fill = PatternFill("solid", fgColor="F1F3F5")
	ws.cell(row=3, column=1).fill = PatternFill(
		"solid", fgColor="F1F3F5",
	)
	ws.cell(row=3, column=total_col).fill = PatternFill(
		"solid", fgColor="F1F3F5",
	)
	ws.row_dimensions[3].height = 32

	# --- Sign-flip + zero-row prune ---
	# Apply FLIP_ROOT_TYPES then compute the keep-set: accounts with any
	# non-zero column, plus every ancestor of such an account. Mirrors
	# the frontend rule in pivot_grid.js _visibleAccounts.
	flipped = {}  # acct_id -> {company: signed_value}
	by_id = {a["id"]: a for a in accounts}
	for acct in accounts:
		flip = -1 if acct.get("root_type") in FLIP_ROOT_TYPES else 1
		row = balances.get(acct["id"]) or {}
		flipped[acct["id"]] = {
			c: float(flt(v)) * flip for c, v in row.items()
		}

	visible_companies = [c for (_t, _n, _col, c) in company_cols]
	keep = set()
	for acct in accounts:
		row = flipped.get(acct["id"]) or {}
		if any((row.get(c) or 0) != 0 for c in visible_companies):
			keep.add(acct["id"])
			cur = acct
			while cur.get("parent"):
				p = by_id.get(cur["parent"])
				if not p:
					break
				keep.add(p["id"])
				cur = p

	# --- Data rows ---
	row_idx = 4
	for acct in accounts:
		if acct["id"] not in keep:
			continue

		depth = int(acct.get("depth") or 0)
		is_group = bool(acct.get("is_group"))
		row_font = group_font if is_group else None
		num_align = Alignment(
			horizontal="right", vertical="center",
			indent=depth * NUMBER_INDENT_PER_DEPTH,
		)

		label = ws.cell(row=row_idx, column=1, value=acct["name"])
		label.alignment = Alignment(
			horizontal="left", vertical="center",
			indent=depth * LABEL_INDENT_PER_DEPTH,
		)
		if row_font:
			label.font = row_font

		flipped_row = flipped.get(acct["id"]) or {}
		row_total = 0.0
		for i, (_tid, _tname, _tcolor, company) in enumerate(company_cols):
			v = float(flipped_row.get(company, 0) or 0)
			row_total += v
			cell = ws.cell(
				row=row_idx, column=2 + i,
				value=v if v != 0 else None,
			)
			cell.number_format = _NUM_FMT
			cell.alignment = num_align
			if row_font:
				cell.font = row_font

		total_cell = ws.cell(
			row=row_idx, column=total_col,
			value=row_total if row_total != 0 else None,
		)
		total_cell.number_format = _NUM_FMT
		total_cell.alignment = num_align
		if row_font:
			total_cell.font = row_font

		row_idx += 1

	# --- Bottom total row (leaf accounts only, matches pivot footer) ---
	leaf_totals = [0.0] * n_cos
	grand = 0.0
	for acct in accounts:
		if acct.get("is_group"):
			continue
		if acct["id"] not in keep:
			continue
		row = flipped.get(acct["id"]) or {}
		for i, (_tid, _tname, _tcolor, company) in enumerate(company_cols):
			v = float(row.get(company, 0) or 0)
			leaf_totals[i] += v
			grand += v

	footer_label = ws.cell(
		row=row_idx, column=1, value="Total (leaf accounts)",
	)
	footer_label.font = group_font
	footer_label.alignment = left_align
	footer_label.fill = PatternFill("solid", fgColor="E9ECEF")
	for i, t in enumerate(leaf_totals):
		c = ws.cell(
			row=row_idx, column=2 + i,
			value=t if t != 0 else None,
		)
		c.number_format = _NUM_FMT
		c.alignment = right_align
		c.font = group_font
		c.fill = PatternFill("solid", fgColor="E9ECEF")
	gc = ws.cell(
		row=row_idx, column=total_col,
		value=grand if grand != 0 else None,
	)
	gc.number_format = _NUM_FMT
	gc.alignment = right_align
	gc.font = group_font
	gc.fill = PatternFill("solid", fgColor="E9ECEF")

	# --- Layout polish: widths + frozen panes ---
	# Deeper rows pull their numbers inward by `depth *
	# NUMBER_INDENT_PER_DEPTH` character widths; without a little
	# extra room a deep account with a wide value like
	# "58,30,28,364.49" would crowd the column. 24 chars fits
	# comfortably even at depth 4.
	ws.column_dimensions["A"].width = 52
	for i in range(n_cos):
		ws.column_dimensions[get_column_letter(2 + i)].width = 24
	ws.column_dimensions[get_column_letter(total_col)].width = 26

	# Freeze the title + two header rows AND the account column.
	ws.freeze_panes = "B4"

	buf = BytesIO()
	wb.save(buf)
	return buf.getvalue()


def _build_empty_xlsx(snapshot_date):
	"""Empty-but-valid xlsx used when the user has no in-scope companies."""
	from openpyxl import Workbook
	from openpyxl.styles import Font

	wb = Workbook()
	ws = wb.active
	ws.title = "Trial Balance"
	cell = ws.cell(
		row=1, column=1,
		value=f"Trial Balance — {snapshot_date.isoformat()}",
	)
	cell.font = Font(bold=True, size=14)
	ws.cell(
		row=3, column=1,
		value="No companies in scope.",
	)
	ws.column_dimensions["A"].width = 60
	buf = BytesIO()
	wb.save(buf)
	return buf.getvalue()


# ---------------------------------------------------------------------------
# Response plumbing
# ---------------------------------------------------------------------------

def _send_xlsx(xlsx_bytes, snapshot_date):
	"""Set Frappe's response so the browser triggers a file download.

	Mirrors gl_drill_v1._set_csv_response: `type=binary` plus an
	explicit filename drives the Content-Disposition header. The
	`.xlsx` extension makes Excel / LibreOffice / Numbers pick the
	right opener.
	"""
	filename = _xlsx_filename(snapshot_date)
	frappe.local.response.filename = filename
	frappe.local.response.filecontent = xlsx_bytes
	frappe.local.response.type = "binary"


def _xlsx_filename(snapshot_date):
	ts = datetime.now().strftime("%H%M%S")
	return f"trial_balance_{snapshot_date.isoformat()}_{ts}.xlsx"


_HEX_RE = re.compile(r"^#?([0-9A-Fa-f]{6})$")


def _hex_to_argb(color):
	"""Convert a `#RRGGBB` trust color to an openpyxl ARGB string."""
	if not color:
		return "FF888888"
	m = _HEX_RE.match(color.strip())
	if not m:
		return "FF888888"
	return f"FF{m.group(1).upper()}"
